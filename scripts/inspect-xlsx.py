"""Inspeciona estrutura da planilha shinova_todos_produtos.xlsx."""
from pathlib import Path
import sys
try:
    from openpyxl import load_workbook
except ImportError:
    print("Pillow está, mas openpyxl não. Instalando...")
    sys.exit(99)

SRC = Path(__file__).resolve().parent.parent / "shinova_todos_produtos.xlsx"

wb = load_workbook(SRC, data_only=True)
print(f"Planilha: {SRC.name}")
print(f"Abas: {wb.sheetnames}")
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"==== Aba: '{sheet_name}' ({ws.max_row} linhas, {ws.max_column} colunas) ====")
    # Cabeçalho
    headers = [cell.value for cell in ws[1]]
    print(f"Colunas: {headers}")
    # Primeiras 5 linhas como exemplo
    print("Primeiras 5 linhas:")
    for row_idx in range(2, min(7, ws.max_row + 1)):
        row_data = {headers[i]: ws.cell(row=row_idx, column=i + 1).value for i in range(len(headers))}
        print(f"  {row_idx}: {row_data}")
    print()
