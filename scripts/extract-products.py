"""
Lê shinova_todos_produtos.xlsx (8 abas = 8 categorias) e gera:
- data/categorias.json (8 categorias mapeadas para PT-BR)
- data/produtos.json   (todos os produtos com galeria de URLs)
- data/produtos-by-categoria.json (índice por slug)

Uso:
    python scripts/extract-products.py
"""
from __future__ import annotations
import json
import re
import sys
import unicodedata
from pathlib import Path
from openpyxl import load_workbook

# UTF-8 friendly em Windows
sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "shinova_todos_produtos.xlsx"
OUT_DIR = ROOT / "data"
OUT_DIR.mkdir(exist_ok=True)

# Mapeamento EN (aba) → PT-BR (categoria)
CATEGORIAS_MAP = {
    "Anesthesia-Monitoring": {
        "slug": "anestesia-monitorizacao",
        "nome": "Anestesia & Monitorização",
        "descricao_curta": "Monitores multiparâmetros, anestesia inalatória, oxímetros e equipamentos para monitorização vital.",
        "numero": "01",
        "ordem": 1,
    },
    "Medical-Imaging": {
        "slug": "imagem-diagnostico",
        "nome": "Imagem & Diagnóstico",
        "descricao_curta": "Ultrassom, raio-X digital, endoscopia e tecnologias de imagem clínica.",
        "numero": "02",
        "ordem": 2,
    },
    "Laboratory-Diagnostics": {
        "slug": "laboratorio-clinico",
        "nome": "Laboratório Clínico",
        "descricao_curta": "Analisadores hematológicos, bioquímicos e centrifugas para diagnóstico laboratorial.",
        "numero": "03",
        "ordem": 3,
    },
    "Treatment-Recovery": {
        "slug": "tratamento-recuperacao",
        "nome": "Tratamento & Recuperação",
        "descricao_curta": "Equipamentos para internação, recuperação cirúrgica e cuidados intensivos.",
        "numero": "04",
        "ordem": 4,
    },
    "Dental-Care": {
        "slug": "odontologia-veterinaria",
        "nome": "Odontologia Veterinária",
        "descricao_curta": "Unidades odontológicas, scalers ultrassônicos e instrumentais para odontologia animal.",
        "numero": "05",
        "ordem": 5,
    },
    "Ophthalmic-Care": {
        "slug": "oftalmologia-veterinaria",
        "nome": "Oftalmologia Veterinária",
        "descricao_curta": "Tonômetros, oftalmoscópios e equipamentos especializados em oftalmologia veterinária.",
        "numero": "06",
        "ordem": 6,
    },
    "Exam-Diagnostics": {
        "slug": "exame-diagnostico",
        "nome": "Exame & Diagnóstico",
        "descricao_curta": "Estetoscópios, otoscópios, balanças e equipamentos para exame clínico de rotina.",
        "numero": "07",
        "ordem": 7,
    },
    "Pet-Grooming": {
        "slug": "pet-grooming-estetica",
        "nome": "Pet Grooming & Estética",
        "descricao_curta": "Mesas hidráulicas, secadores, banheiras e equipamentos para banho e tosa profissional.",
        "numero": "08",
        "ordem": 8,
    },
}


def slugify(text: str) -> str:
    """Converte texto em slug URL-safe (sem acento, kebab-case)."""
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text).encode("ASCII", "ignore").decode("ASCII")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-+", "-", text)
    return text.strip("-")


def parse_imagens(raw: str | None) -> list[str]:
    if not raw:
        return []
    # separadores possíveis: | , \n
    parts = re.split(r"\s*\|\s*|\n", str(raw))
    urls = [p.strip() for p in parts if p.strip().startswith("http")]
    return urls


def normalize_text(s: str | None) -> str | None:
    if not s:
        return None
    s = str(s).replace("\r\n", "\n").replace("\r", "\n")
    # remove múltiplas quebras
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip() or None


def parse_especificacoes(raw: str | None) -> list[dict[str, str]]:
    """
    Tenta extrair lista de {chave, valor}. Se vier em formato livre, devolve
    como item único {chave: 'Especificações', valor: texto}.
    """
    if not raw:
        return []
    text = normalize_text(raw) or ""
    specs: list[dict[str, str]] = []

    # Padrão chave: valor por linha
    for line in text.split("\n"):
        line = line.strip()
        if not line:
            continue
        m = re.match(r"^([^:]{2,40}):\s*(.+)$", line)
        if m:
            specs.append({"chave": m.group(1).strip(), "valor": m.group(2).strip()})

    if not specs:
        # fallback — devolve texto bruto
        return [{"chave": "Especificações", "valor": text}]
    return specs


def extract() -> None:
    print(f"Lendo {SRC.name}")
    wb = load_workbook(SRC, data_only=True)
    if not SRC.exists():
        raise FileNotFoundError(SRC)

    categorias_out: list[dict] = []
    produtos_out: list[dict] = []
    slugs_seen: set[str] = set()

    for aba in wb.sheetnames:
        if aba not in CATEGORIAS_MAP:
            print(f"⚠️ aba ignorada (sem mapeamento): {aba}")
            continue

        cat_meta = CATEGORIAS_MAP[aba]
        cat_id = cat_meta["slug"]
        categorias_out.append({
            "id": cat_id,
            "slug": cat_meta["slug"],
            "nome": cat_meta["nome"],
            "descricao_curta": cat_meta["descricao_curta"],
            "numero": cat_meta["numero"],
            "ordem": cat_meta["ordem"],
            "destaque": cat_meta["ordem"] <= 4,
            "fonte_aba": aba,
        })

        ws = wb[aba]
        headers = [cell.value for cell in ws[1]]
        if not headers:
            continue

        col_idx = {h: i for i, h in enumerate(headers) if h}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[col_idx.get("Nome", 0)]:
                continue

            nome = str(row[col_idx["Nome"]]).strip()
            modelo = str(row[col_idx.get("Modelo", 1)] or "").strip()
            subcategoria = row[col_idx.get("Subcategoria", 3)]
            descricao = row[col_idx.get("Descricao", 4)]
            especificacoes = row[col_idx.get("Especificacoes", 5)]
            configuracoes = row[col_idx.get("Configuracoes", 6)]
            imagens_raw = row[col_idx.get("Imagens (URLs)", 11)] if "Imagens (URLs)" in col_idx else None
            url_produto = row[col_idx.get("URL Produto", 12)] if "URL Produto" in col_idx else None

            # slug a partir de modelo+nome, garante unicidade
            base = modelo or nome
            slug = slugify(f"{base}")
            if not slug:
                continue
            n = 1
            unique = slug
            while unique in slugs_seen:
                n += 1
                unique = f"{slug}-{n}"
            slugs_seen.add(unique)

            imagens = parse_imagens(str(imagens_raw) if imagens_raw else None)
            galeria = [{"url": url, "ordem": i, "alt": f"{nome} - imagem {i + 1}"} for i, url in enumerate(imagens)]

            produtos_out.append({
                "id": unique,
                "slug": unique,
                "categoria_id": cat_id,
                "categoria_slug": cat_meta["slug"],
                "categoria_nome": cat_meta["nome"],
                "modelo": modelo or None,
                "nome": nome,
                "marca": "SHINOVA",
                "subcategoria": str(subcategoria).strip() if subcategoria else None,
                "descricao_curta": (normalize_text(descricao) or "").split("\n")[0][:200] if descricao else None,
                "descricao_longa": normalize_text(descricao),
                "especificacoes": parse_especificacoes(str(especificacoes) if especificacoes else None),
                "configuracoes": normalize_text(str(configuracoes) if configuracoes else None),
                "imagem_principal": imagens[0] if imagens else None,
                "galeria": galeria,
                "url_fabricante": str(url_produto) if url_produto else None,
                "destaque": False,
                "publicado": True,
            })

    # marcar alguns produtos como destaque (primeiros de cada categoria)
    por_categoria: dict[str, list] = {}
    for p in produtos_out:
        por_categoria.setdefault(p["categoria_id"], []).append(p)
    for cat_id, prods in por_categoria.items():
        for p in prods[:1]:
            p["destaque"] = True

    # salva
    (OUT_DIR / "categorias.json").write_text(
        json.dumps(categorias_out, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (OUT_DIR / "produtos.json").write_text(
        json.dumps(produtos_out, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    indice = {
        cat_id: [p["slug"] for p in prods] for cat_id, prods in por_categoria.items()
    }
    (OUT_DIR / "produtos-by-categoria.json").write_text(
        json.dumps(indice, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(f"\n✅ {len(categorias_out)} categorias salvas")
    print(f"✅ {len(produtos_out)} produtos salvos")
    print()
    print("Resumo por categoria:")
    for cat in categorias_out:
        n = len(por_categoria.get(cat["id"], []))
        print(f"  - {cat['numero']} {cat['nome']:<32} {n:>3} produtos")

    print(f"\nArquivos gerados em {OUT_DIR}/")


if __name__ == "__main__":
    extract()
